import openpyxl

# Crear un nuevo libro de Excel
workbook = openpyxl.Workbook()

# Seleccionar la hoja activa y establecer su título
sheet = workbook.active
sheet.title = "Lista de Supermercados"

# Agregar encabezados
sheet["A1"] = "Supermercado" 
sheet["B1"] = "Producto"
sheet["C1"] = "Precio"

# Lista de supermercados con nombre de producto y precio
supermarkets = [
    ("Supermercado A", "Manzanas", 2.99),
    ("Supermercado A", "Peras", 3.49),
    ("Supermercado B", "Naranjas", 2.79),
    ("Supermercado B", "Kiwi", 3.29),
    ("Supermercado C", "Jocote", 2.50),
    ("Supermercado c", "Pepino", 4.54),
    ("Supermercado B", "Zanahoria", 2.3),
    
]

# Agregar datos a la hoja
for row, data in enumerate(supermarkets, start=2):
    sheet.cell(row=row, column=1, value=data[0])
    sheet.cell(row=row, column=2, value=data[1])
    sheet.cell(row=row, column=3, value=data[2])

# Guardar el archivo Excel
workbook.save("lista_supermercados.xlsx")

# Cerrar el archivo Excel
workbook.close()